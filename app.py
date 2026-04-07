import streamlit as st
import tempfile
import os
import io
from main import process_candidates


# PAGE CONFIG

st.set_page_config(
    page_title="ATS Resume Screener",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# CUSTOM CSS

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Sans:wght@300;400;500;600&display=swap');

/* Reset & Base */
html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
}

.stApp {
    background-color: #F7F5F0;
}

/* Hide streamlit default elements */
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 2.5rem 3rem; max-width: 1100px; }

/* Header */
.ats-header {
    border-bottom: 2px solid #1C1C1C;
    padding-bottom: 1.5rem;
    margin-bottom: 2.5rem;
}
.ats-title {
    font-family: 'DM Serif Display', serif;
    font-size: 2.6rem;
    color: #1C1C1C;
    letter-spacing: -0.5px;
    margin: 0;
    line-height: 1.1;
}
.ats-subtitle {
    font-size: 0.92rem;
    color: #777;
    margin-top: 0.4rem;
    font-weight: 300;
    letter-spacing: 0.5px;
    text-transform: uppercase;
}

/* Section labels */
.section-label {
    font-size: 0.72rem;
    font-weight: 600;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    color: #999;
    margin-bottom: 0.5rem;
}

/* Input styling */
.stTextArea textarea {
    background: #fff !important;
    border: 1.5px solid #E0DDD8 !important;
    border-radius: 4px !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.9rem !important;
    color: #1C1C1C !important;
    padding: 1rem !important;
    transition: border-color 0.2s;
}
.stTextArea textarea:focus {
    border-color: #1C1C1C !important;
    box-shadow: none !important;
}

/* File uploader */
[data-testid="stFileUploader"] {
    background: #fff;
    border: 1.5px dashed #C8C4BE;
    border-radius: 4px;
    padding: 1rem;
}
[data-testid="stFileUploader"]:hover {
    border-color: #1C1C1C;
}

/* Button */
.stButton > button {
    background: #1C1C1C !important;
    color: #F7F5F0 !important;
    border: none !important;
    border-radius: 3px !important;
    padding: 0.65rem 2.2rem !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.88rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.5px !important;
    cursor: pointer !important;
    transition: background 0.2s !important;
}
.stButton > button:hover {
    background: #333 !important;
}

/* Download button */
.stDownloadButton > button {
    background: transparent !important;
    color: #1C1C1C !important;
    border: 1.5px solid #1C1C1C !important;
    border-radius: 3px !important;
    padding: 0.5rem 1.4rem !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.82rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.3px !important;
}
.stDownloadButton > button:hover {
    background: #1C1C1C !important;
    color: #F7F5F0 !important;
}

/* Candidate card */
.candidate-card {
    background: #fff;
    border: 1.5px solid #E0DDD8;
    border-radius: 6px;
    padding: 1.8rem 2rem;
    margin-bottom: 1.2rem;
    position: relative;
}
.candidate-rank {
    position: absolute;
    top: 1.6rem;
    right: 1.8rem;
    font-size: 0.72rem;
    font-weight: 600;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    color: #999;
}
.candidate-name {
    font-family: 'DM Serif Display', serif;
    font-size: 1.5rem;
    color: #1C1C1C;
    margin: 0 0 0.2rem 0;
}
.score-bar-wrap {
    background: #F0EDE8;
    border-radius: 2px;
    height: 4px;
    margin: 0.8rem 0 1.2rem 0;
    width: 100%;
}
.score-bar-fill {
    height: 4px;
    border-radius: 2px;
    background: #1C1C1C;
    transition: width 0.6s ease;
}
.score-number {
    font-size: 0.82rem;
    font-weight: 600;
    color: #1C1C1C;
    letter-spacing: 0.3px;
}
.meta-grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 1rem;
    margin: 1rem 0;
}
.meta-item-label {
    font-size: 0.7rem;
    font-weight: 600;
    letter-spacing: 1.2px;
    text-transform: uppercase;
    color: #999;
    margin-bottom: 0.25rem;
}
.meta-item-value {
    font-size: 0.88rem;
    color: #1C1C1C;
    font-weight: 400;
    line-height: 1.4;
}
.summary-box {
    background: #F7F5F0;
    border-left: 3px solid #1C1C1C;
    padding: 0.9rem 1.1rem;
    margin: 1rem 0;
    border-radius: 0 3px 3px 0;
}
.summary-text {
    font-size: 0.88rem;
    color: #444;
    line-height: 1.65;
    font-style: italic;
}
.skills-wrap {
    display: flex;
    flex-wrap: wrap;
    gap: 0.4rem;
    margin-top: 0.5rem;
}
.skill-tag {
    background: #F0EDE8;
    color: #1C1C1C;
    font-size: 0.75rem;
    font-weight: 500;
    padding: 0.3rem 0.7rem;
    border-radius: 2px;
    letter-spacing: 0.2px;
}
.divider {
    border: none;
    border-top: 1px solid #E0DDD8;
    margin: 1.2rem 0;
}
.result-header {
    display: flex;
    align-items: baseline;
    justify-content: space-between;
    margin-bottom: 1.5rem;
    padding-bottom: 0.8rem;
    border-bottom: 1px solid #E0DDD8;
}
.result-count {
    font-family: 'DM Serif Display', serif;
    font-size: 1.1rem;
    color: #1C1C1C;
}
</style>
""", unsafe_allow_html=True)

# EXCEL EXPORT

def generate_excel(results):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidate Report"

        # Header row styling
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill("solid", start_color="1C1C1C")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        headers = ["Rank", "Name", "Match Score (%)", "Experience", "Education", "Skills", "Summary"]
        col_widths = [6, 22, 16, 28, 30, 45, 60]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 30

        # Data rows
        alt_fill = PatternFill("solid", start_color="F7F5F0")
        data_font = Font(name="Arial", size=10)
        data_align = Alignment(vertical="top", wrap_text=True)

        for row_idx, res in enumerate(results, start=2):
            skills_str = ", ".join(res.get("skills", [])) if res.get("skills") else "N/A"
            row_data = [
                row_idx - 1,
                res.get("name", "Unknown"),
                res.get("score", 0),
                res.get("experience", "N/A"),
                res.get("education", "N/A"),
                skills_str,
                res.get("summary", "N/A"),
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

            ws.row_dimensions[row_idx].height = 60

        # Score column number format
        for row_idx in range(2, len(results) + 2):
            ws.cell(row=row_idx, column=3).number_format = '0.00"%"'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    except Exception as e:
        print(f"Excel generation error: {e}")
        return None

# UI

st.markdown("""
<div class="ats-header">
    <p class="ats-title">Resume Screening</p>
    <p class="ats-subtitle">Applicant Tracking System &nbsp;&middot;&nbsp;</p>
</div>
""", unsafe_allow_html=True)

col1, col2 = st.columns([3, 2], gap="large")

with col1:
    st.markdown('<p class="section-label">Job Description</p>', unsafe_allow_html=True)
    jd = st.text_area(
        label="jd",
        label_visibility="collapsed",
        placeholder="Paste the full job description here...",
        height=220
    )

with col2:
    st.markdown('<p class="section-label">Upload Resumes</p>', unsafe_allow_html=True)
    uploaded_files = st.file_uploader(
        label="resumes",
        label_visibility="collapsed",
        type=["pdf"],
        accept_multiple_files=True
    )
    if uploaded_files:
        st.markdown(
            f'<p style="font-size:0.8rem;color:#999;margin-top:0.3rem;">'
            f'{len(uploaded_files)} file{"s" if len(uploaded_files) > 1 else ""} selected</p>',
            unsafe_allow_html=True
        )

st.markdown("<br>", unsafe_allow_html=True)
run_btn = st.button("Run Analysis")

# RESULTS

if run_btn:
    if not jd or not jd.strip():
        st.warning("Please enter a job description.")
    elif not uploaded_files:
        st.warning("Please upload at least one resume.")
    else:
        file_paths = []
        for file in uploaded_files:
            temp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
            temp.write(file.read())
            temp.flush()
            temp.close()
            file_paths.append(temp.name)

        with st.spinner("Analyzing candidates..."):
            results = process_candidates(file_paths, jd)

        for path in file_paths:
            try:
                os.unlink(path)
            except:
                pass

        if not results:
            st.info("No results returned. Check terminal for errors.")
        else:
            st.markdown("<br>", unsafe_allow_html=True)
            dl_col, count_col = st.columns([1, 3])
            with count_col:
                st.markdown(
                    f'<div class="result-count">{len(results)} candidate{"s" if len(results) > 1 else ""} screened</div>',
                    unsafe_allow_html=True
                )
            with dl_col:
                excel_data = generate_excel(results)
                if excel_data:
                    st.download_button(
                        label="Download Report (.xlsx)",
                        data=excel_data,
                        file_name="ats_candidate_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

            st.markdown("<br>", unsafe_allow_html=True)

            for rank, res in enumerate(results, start=1):
                name = res.get("name", "Unknown")
                score = res.get("score", 0)
                experience = res.get("experience", "N/A")
                education = res.get("education", "N/A")
                skills = res.get("skills", [])
                summary = res.get("summary", "N/A")

                skills_html = "".join(
                    f'<span class="skill-tag">{s}</span>'
                    for s in skills
                ) if skills else '<span style="color:#999;font-size:0.82rem;">No skills extracted</span>'

                st.markdown(f"""
<div class="candidate-card">
    <span class="candidate-rank">#{rank}</span>
    <p class="candidate-name">{name}</p>
    <span class="score-number">{score}% match</span>
    <div class="score-bar-wrap">
        <div class="score-bar-fill" style="width:{score}%"></div>
    </div>
    <div class="meta-grid">
        <div>
            <div class="meta-item-label">Experience</div>
            <div class="meta-item-value">{experience}</div>
        </div>
        <div>
            <div class="meta-item-label">Education</div>
            <div class="meta-item-value">{education}</div>
        </div>
    </div>
    <hr class="divider">
    <div class="meta-item-label">Professional Summary</div>
    <div class="summary-box">
        <p class="summary-text">{summary}</p>
    </div>
    <hr class="divider">
    <div class="meta-item-label">Skills</div>
    <div class="skills-wrap">{skills_html}</div>
</div>
""", unsafe_allow_html=True)